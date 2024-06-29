import os
import tensorflow as tf
from keras.models import Sequential, load_model
from keras.layers import Flatten, Dense, Conv2D, MaxPooling2D, Dropout
from keras.layers import BatchNormalization
from keras.layers.core import Activation
from keras.preprocessing.image import ImageDataGenerator
from keras.utils import to_categorical
import cv2
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from pynvml import *
from keras.callbacks import TensorBoard, ModelCheckpoint
import datetime
import psutil
import time
from tensorflow import compat
from keras import backend as K
import shutil
import openpyxl

class AlexNet:

    # def __init__(self):
    #     # Inicjalizacja NVML
    #     nvmlInit()
    #     self.device_count = nvmlDeviceGetCount()
    #     self.gpu_usage_data = []
    
    # def __del__(self):
    #     # Zamykanie NVML przy zniszczeniu obiektu
    #     nvmlShutdown()

    def display_history(self, history, training_time, model_name, arch_name, accuracy, loss):
        # df = pd.DataFrame(history)
        # ax = df.plot(figsize=(8, 5))
        # plt.grid(True)
        # # plt.gca().set_ylim(0, 1)

        # # Dodanie tekstu pod wykresem
        # plt.subplots_adjust(bottom=0.2)  # Adjust bottom to make space for text
        # plt.text(0.5, -0.15, f'Czas treningu: {training_time:.2f} sekund', color='red', 
        #          ha='center', va='top', transform=ax.transAxes, fontsize=12)

        # plt.savefig('out/history.png')
        # plt.show()

        df = pd.DataFrame(history)
        
        fig, ax1 = plt.subplots(figsize=(8, 5))

        # Skala po lewej stronie dla accuracy
        ax1.plot(df.index, df['accuracy'], 'b-', label='Accuracy')
        ax1.plot(df.index, df['val_accuracy'], 'g-', label='Validation Accuracy')
        ax1.set_xlabel('Epoch')
        ax1.set_ylabel('Accuracy', color='b')
        ax1.tick_params(axis='y', labelcolor='b')
        ax1.legend(loc='upper left')

        parts = model_name.split('__')
        model_params = {}
        for part in parts:
            key_value = part.split('-', 1)  # Split only on the first dash
            key = key_value[0]
            value = key_value[1]
            if key in model_params:
                model_params[key] += '-' + value
            else:
                model_params[key] = value
        
        title_part = f"Batch size: {model_params['bs']}, Activation function: {model_params['af']}, Learning rate: {model_params['lr']}"

        ax1.set_title('Training History + GPU Usage:\n {} - {}'.format(arch_name, title_part))

        # Tworzenie drugiej osi y dla loss
        ax2 = ax1.twinx()
        ax2.plot(df.index, df['loss'], 'r-', label='Loss')
        ax2.plot(df.index, df['val_loss'], 'm-', label='Validation Loss')
        ax2.set_ylabel('Loss', color='r')
        ax2.tick_params(axis='y', labelcolor='r')
        ax2.legend(loc='upper right')

        fig.tight_layout()  # Dostosowanie layoutu, żeby elementy się nie nakładały
        ax1.grid(True)

        # Dodanie tekstu pod wykresem
        plt.subplots_adjust(bottom=0.25)  # Adjust bottom to make space for text
        plt.text(0.5, -0.30, f'Czas treningu: {training_time:.2f} sekund', color='red', 
                 ha='center', va='top', transform=ax1.transAxes, fontsize=12)
        
        # Wyświetlenie Accuracy i Loss
        plt.text(0.5, -0.16, f'Accuracy: {accuracy:.4f}', color='green', ha='center', va='top', transform=ax1.transAxes, fontsize=12)
        plt.text(0.5, -0.23, f'Loss: {loss:.4f}', color='blue', ha='center', va='top', transform=ax1.transAxes, fontsize=12)

        disp_dir = os.path.join("out", arch_name)
        if not os.path.exists(disp_dir):
            os.makedirs(disp_dir)
        plt.savefig(disp_dir + '/' + model_name + '__' + datetime.datetime.now().strftime("%Y%m%d-%H%M%S") + '.png')

        # plt.show()

    def display_combined_history(self, history, gpu_usage_data, training_time, model_name, arch_name, accuracy, loss):
        # Tworzenie wykresu łączonego dla historii trenowania i danych GPU
        history_df = pd.DataFrame(history)
        gpu_usage_df = pd.DataFrame(gpu_usage_data)
        
        fig, ax1 = plt.subplots(figsize=(8, 5))

        ax1.set_xlabel('Epoch')
        ax1.set_ylabel('Accuracy / Loss')
        ax1.plot(history_df.index, history_df['accuracy'], 'b-', label='Accuracy')
        ax1.plot(history_df.index, history_df['val_accuracy'], 'g-', label='Validation Accuracy')
        ax1.plot(history_df.index, history_df['loss'], 'r-', label='Loss')
        ax1.plot(history_df.index, history_df['val_loss'], 'm-', label='Validation Loss')
        ax1.legend(loc='upper left')

        parts = model_name.split('__')
        model_params = {}
        for part in parts:
            key_value = part.split('-', 1)  # Split only on the first dash
            key = key_value[0]
            value = key_value[1]
            if key in model_params:
                model_params[key] += '-' + value
            else:
                model_params[key] = value
        
        title_part = f"Batch size: {model_params['bs']}, Activation function: {model_params['af']}, Learning rate: {model_params['lr']}"

        ax1.set_title('Training History + GPU Usage:\n {} - {}'.format(arch_name, title_part))
        
        ax2 = ax1.twinx()
        ax2.set_ylabel('GPU Utilization (%) / Memory Utilization (%)')
        ax2.plot(gpu_usage_df.index, gpu_usage_df['gpu_utilization'], 'k-', label='GPU Utilization')
        ax2.plot(gpu_usage_df.index, gpu_usage_df['memory_utilization'], 'c-', label='Memory Utilization')
        ax2.legend(loc='upper right')

        fig.tight_layout()
        ax1.grid(True)
        plt.grid(True)
        # Dodanie tekstu pod wykresem
        plt.subplots_adjust(bottom=0.25)  # Adjust bottom to make space for text
        plt.text(0.5, -0.30, f'Czas treningu: {training_time:.2f} sekund', color='red', 
                 ha='center', va='top', transform=ax1.transAxes, fontsize=12)
        
        # Wyświetlenie Accuracy i Loss
        plt.text(0.5, -0.16, f'Accuracy: {accuracy:.4f}', color='green', ha='center', va='top', transform=ax1.transAxes, fontsize=12)
        plt.text(0.5, -0.23, f'Loss: {loss:.4f}', color='blue', ha='center', va='top', transform=ax1.transAxes, fontsize=12)

        disp_dir = os.path.join("out", arch_name)
        if not os.path.exists(disp_dir):
            os.makedirs(disp_dir)
        plt.savefig(disp_dir + '/' + 'combined__' + model_name + '__' + datetime.datetime.now().strftime("%Y%m%d-%H%M%S") + '.png')

        # plt.show()

    # Function to find the next available row in a given column
    def find_next_available_row(self, ws, column):
        row = 3
        while ws[f"{column}{row}"].value is not None:
            row += 1
        return row

    
    def train_model(self, model_name, arch_name, compile_optimizer, compile_loss, fit_epochs, fit_batch_size, activation_function, log_custom_dir=''):

        model = Sequential();

        # # AlexNet Implementation
        # model.add(Conv2D(filters=96, kernel_size=(11,11), strides=(4,4), input_shape=(227, 227, 3)))
        # model.add(Activation("relu"))
        # model.add(MaxPooling2D(pool_size=(3,3), strides=(2,2)))

        # model.add(Conv2D(filters=256, kernel_size=(5,5), strides=(1,1), padding='same'))
        # model.add(Activation("relu"))
        # model.add(MaxPooling2D(pool_size=(3,3), strides=(2,2)))

        # model.add(Conv2D(filters=384, kernel_size=(3,3), strides=(1,1), padding='same'))
        # model.add(Activation("relu"))
        # model.add(Conv2D(filters=384, kernel_size=(3,3), strides=(1,1), padding='same'))
        # model.add(Activation("relu"))
        # model.add(Conv2D(filters=256, kernel_size=(3,3), strides=(1,1), padding='same'))
        # model.add(Activation("relu"))
        # model.add(MaxPooling2D(pool_size=(3,3), strides=(2,2)))

        # model.add(Flatten())
        # model.add(Dense(units=4096))
        # model.add(Activation("relu"))
        # model.add(Dropout(0.5))
        # model.add(Dense(units=4096))
        # model.add(Activation("relu"))
        # model.add(Dropout(0.5))
        # model.add(Dense(units=10)) # 10 - liczba klas w ImageNet
        # model.add(Activation("softmax"))

        # Warstwa 1: Conv2D -> Activation -> BatchNormalization -> MaxPooling2D
        model.add(Conv2D(filters=96, kernel_size=(11,11), strides=(4,4), input_shape=(227, 227, 3), activation=activation_function))
        model.add(BatchNormalization())
        model.add(MaxPooling2D(pool_size=(3,3), strides=(2,2)))

        # Warstwa 2: Conv2D -> Activation -> BatchNormalization -> MaxPooling2D
        model.add(Conv2D(filters=256, kernel_size=(5,5), padding='valid', activation=activation_function))
        model.add(BatchNormalization())
        model.add(MaxPooling2D(pool_size=(3,3), strides=(2,2)))

        # Warstwa 3: Conv2D -> Activation -> BatchNormalization
        model.add(Conv2D(filters=384, kernel_size=(3,3), padding='valid', activation=activation_function))
        model.add(BatchNormalization())

        # Warstwa 4: Conv2D -> Activation -> BatchNormalization
        model.add(Conv2D(filters=384, kernel_size=(3,3), padding='valid', activation=activation_function))
        model.add(BatchNormalization())

        # Warstwa 5: Conv2D -> Activation -> BatchNormalization -> MaxPooling2D
        model.add(Conv2D(filters=256, kernel_size=(3,3), padding='valid', activation=activation_function))
        model.add(BatchNormalization())
        model.add(MaxPooling2D(pool_size=(3,3), strides=(2,2)))

        # Warstwa 6: Flatten
        model.add(Flatten())

        # Warstwa 7: Dense -> Activation -> Dropout
        model.add(Dense(4096, activation=activation_function))
        model.add(Dropout(0.5))

        # Warstwa 8: Dense -> Activation -> Dropout
        model.add(Dense(4096, activation=activation_function))
        model.add(Dropout(0.5))

        # Warstwa wyjściowa: Dense -> Activation
        num_classes = 10  # liczba klas
        model.add(Dense(num_classes, activation='softmax'))

        # Wyświetlenie podsumowania modelu
        model.summary()


        model.compile(optimizer=compile_optimizer, loss=compile_loss, metrics=['accuracy'])

        train_datagen = ImageDataGenerator(
            rescale=1./255
        )
        test_datagen = ImageDataGenerator(
            rescale=1./255
        )

        train_generator = train_datagen.flow_from_directory(
            'DANE/etap1/raw-img',
            target_size=(227, 227),
            batch_size=fit_batch_size,
            class_mode='categorical')

        test_generator = test_datagen.flow_from_directory(
            'DANE/etap1/test',
            target_size=(227, 227),
            batch_size=fit_batch_size,
            class_mode='categorical')
        
        # Konfiguracja TensorBoard
        log_dir = os.path.join(log_custom_dir, "logs", "fit", arch_name, model_name + '__' + datetime.datetime.now().strftime("%Y%m%d-%H%M%S"))
        tensorboard_callback = TensorBoard(log_dir=log_dir, histogram_freq=4)

        # Callbacks definitions
        system_usage_logger = SystemUsageLogger(log_dir=log_dir, log_frequency=4)
        image_callback = TensorBoardImageCallback(log_dir, train_generator, test_generator, log_frequency=4)
        log_gpu_usage_callback = GPUUsageLogger()
        profiler_callback = ProfilerCallback(log_dir=log_dir, log_frequency=8)
        measuring_time = MeasuringTime()

        model_file_path = os.path.join('models', arch_name)
        if not os.path.exists(model_file_path):
            os.makedirs(model_file_path)
        ckpt_file_path_with_name = os.path.join(model_file_path, 'model__' + model_name + '__' + datetime.datetime.now().strftime("%Y%m%d-%H%M%S") + '.ckpt')
        model_checkpoint_callback = ModelCheckpoint(
            filepath=ckpt_file_path_with_name,  # Ścieżka do pliku zapisu modelu
            save_best_only=True,       # Zapisuj tylko najlepszy model
            monitor='val_accuracy',    # Monitorowana wartość (dokładność walidacyjna)
            mode='max',                # Tryb maksymalizacji wartości (dla 'accuracy')
            verbose=1                  # Wyświetl komunikaty o zapisie modelu
        )


        start_time = time.time()

        history = model.fit(train_generator, 
                                steps_per_epoch=train_generator.samples // fit_batch_size, 
                                epochs=fit_epochs, 
                                validation_data=test_generator, 
                                callbacks = [measuring_time, 
                                             log_gpu_usage_callback, 
                                             system_usage_logger, 
                                             tensorboard_callback, 
                                             image_callback, 
                                             profiler_callback,
                                             model_checkpoint_callback])
        
        stop_time = time.time()

        # Pobieranie danych z `gpu_logger` po zakończeniu trenowania
        gpu_usage_data = log_gpu_usage_callback.on_train_end()

        training_time = measuring_time.on_train_end()

        model = load_model(ckpt_file_path_with_name)
        model_file_path_with_name = os.path.join(model_file_path, 'model__' + model_name + '__' + datetime.datetime.now().strftime("%Y%m%d-%H%M%S") + '.keras')
        model.save(model_file_path_with_name)
        model = load_model(model_file_path_with_name)

        score = model.evaluate(test_generator, verbose=0)
        print('Test loss:', score[0])
        print('Test accuracy:', score[1])
        print("Total time: {}".format(stop_time-start_time))

        # Obliczanie średnich wartości zużycia GPU oraz pamięci
        total_gpu_utilization = sum(data['gpu_utilization'] for data in gpu_usage_data)
        total_memory_utilization = sum(data['memory_utilization'] for data in gpu_usage_data)
        data_len = len(gpu_usage_data)

        average_gpu_utilization = total_gpu_utilization / data_len if data_len > 0 else 0
        average_memory_utilization = total_memory_utilization / data_len if data_len > 0 else 0

        print(f"Average GPU Utilization: {average_gpu_utilization}%")
        print(f"Average Memory Utilization: {average_memory_utilization}%")

        # Obliczanie średnich wartości zużycia RAM oraz CPU
        used_ram_data = system_usage_logger.used_ram_data
        cpu_usage_data = system_usage_logger.cpu_usage_data

        average_used_ram = sum(used_ram_data) / len(used_ram_data) if used_ram_data else 0
        average_cpu_usage = sum(cpu_usage_data) / len(cpu_usage_data) if cpu_usage_data else 0

        print(f"Average Used RAM: {average_used_ram}%")
        print(f"Average CPU Usage: {average_cpu_usage}%")

        # Path to the Excel file
        excel_file_path = os.path.join("out", arch_name, 'Excels')
        if not os.path.exists(excel_file_path):
            os.makedirs(excel_file_path)

        excel_file_path_with_name = os.path.join(excel_file_path, 'excel_name' + '.xlsx')

        # Load or create a new Excel file
        if os.path.exists(excel_file_path_with_name):
            wb = openpyxl.load_workbook(excel_file_path_with_name)
        else:
            wb = openpyxl.Workbook()

        # Select the active sheet
        ws = wb.active

        # Find the next available rows for each column
        row_accuracy = self.find_next_available_row(ws, 'B')
        row_loss = self.find_next_available_row(ws, 'C')
        row_time = self.find_next_available_row(ws, 'D')
        row_gpu_util = self.find_next_available_row(ws, 'E')
        row_mem_util = self.find_next_available_row(ws, 'F')
        row_ram_usage = self.find_next_available_row(ws, 'G')
        row_cpu_usage = self.find_next_available_row(ws, 'H')

        # Ensure rows align (take the maximum row number to avoid overwriting)
        next_row = max(row_accuracy, row_loss, row_time, row_gpu_util, row_mem_util, row_ram_usage, row_cpu_usage)

        # Write the data to the next available row
        ws[f"B{next_row}"] = score[1]
        ws[f"C{next_row}"] = score[0]
        ws[f"D{next_row}"] = training_time
        ws[f"E{next_row}"] = average_gpu_utilization
        ws[f"F{next_row}"] = average_memory_utilization
        ws[f"G{next_row}"] = average_used_ram
        ws[f"H{next_row}"] = average_cpu_usage

        # ws[f"C{next_row}"] = f"{score[0]:.4f}".replace(',', '.')
        # ws[f"D{next_row}"] = f"{training_time:.2f}".replace(',', '.')
        # ws[f"E{next_row}"] = f"{average_gpu_utilization:.2f}".replace(',', '.')
        # ws[f"F{next_row}"] = f"{average_memory_utilization:.2f}".replace(',', '.')
        # ws[f"G{next_row}"] = f"{average_used_ram:.2f}".replace(',', '.')
        # ws[f"H{next_row}"] = f"{average_cpu_usage:.2f}".replace(',', '.')

        # Add the formula in column I starting from row 3 down to the current row
        for row in range(3, next_row + 1):
            formula = f"=ABS(- (3 * (1 - B{row})) - (2 * C{row}) - (2 * (D{row} / MAX(D$3:D$1048576))) - (1 * (E{row} / 100)) - (1 * (F{row} / 100)) - (1 * (G{row} / 100)) - (1 * (H{row} / 100)))"
            ws[f"I{row}"].value = formula
        # Save the workbook
        wb.save(excel_file_path_with_name)

        # Wyświetlenie historii trenowania oraz danych dotyczących użycia GPU
        self.display_history(history.history, training_time, model_name, arch_name, score[1], score[0])
        self.display_combined_history(history.history, gpu_usage_data, training_time, model_name, arch_name, score[1], score[0])

        # # Czyszczenie sesji Keras, aby zwolnić pamięć
        K.clear_session()
        compat.v1.reset_default_graph()

        shutil.rmtree(ckpt_file_path_with_name)

        del model, score, history, train_datagen, test_datagen, train_generator, test_generator, model_checkpoint_callback, compile_optimizer



class TensorBoardImageCallback(tf.keras.callbacks.Callback):
    def __init__(self, log_dir, train_data, test_data, log_frequency=1):
        super().__init__()
        self.log_dir = log_dir
        self.train_data = train_data
        self.test_data = test_data
        self.writer = tf.summary.create_file_writer(os.path.join(log_dir, 'images'))
        self.log_frequency = log_frequency

    def on_epoch_end(self, epoch, logs=None):
        if epoch % self.log_frequency == 0:
            # Wybierz kilka przykładów z danych treningowych i walidacyjnych
            train_images, _ = next(self.train_data)
            test_images, _ = next(self.test_data)
            
            # Przygotuj obrazy do zapisu
            with self.writer.as_default():
                tf.summary.image('Training Images', train_images, step=epoch)
                tf.summary.image('Test Images', test_images, step=epoch)

class SystemUsageLogger(tf.keras.callbacks.Callback):
    def __init__(self, log_dir, log_frequency=3):
        super(SystemUsageLogger, self).__init__()
        self.log_dir = log_dir
        self.file_writer = tf.summary.create_file_writer(log_dir)
        self.log_frequency = log_frequency
        self.used_ram_data = []
        self.cpu_usage_data = []
    
    def on_epoch_end(self, epoch, logs=None):
        if epoch % self.log_frequency == 0:    
            # Log system usage
            memory_info = psutil.virtual_memory()
            used_ram = memory_info.used / (1024 ** 3)  # Convert to GB
            used_ram_percent = memory_info.percent
            available_ram = memory_info.available / (1024 ** 3)  # Convert to GB
            cpu_usage = psutil.cpu_percent()

            self.used_ram_data.append(used_ram_percent)
            self.cpu_usage_data.append(cpu_usage)

            with self.file_writer.as_default():
                tf.summary.scalar('used_ram', used_ram, step=epoch)
                tf.summary.scalar('available_ram', available_ram, step=epoch)
                tf.summary.scalar('cpu_usage', cpu_usage, step=epoch)

class GPUUsageLogger(tf.keras.callbacks.Callback):
    def __init__(self):
        super(GPUUsageLogger, self).__init__()
        nvmlInit()
        self.device_count = nvmlDeviceGetCount()
        self.gpu_usage_data = []

    def __del__(self):
        # Zamykanie NVML przy zniszczeniu obiektu
        nvmlShutdown()

    def get_gpu_usage(self):
        gpu_usages = []
        for i in range(self.device_count):
            handle = nvmlDeviceGetHandleByIndex(i)
            info = nvmlDeviceGetMemoryInfo(handle)
            utilization = nvmlDeviceGetUtilizationRates(handle)
            gpu_usages.append({
                'gpu': i,
                'memory_total': info.total,
                'memory_free': info.free,
                'memory_used': info.used,
                'gpu_utilization': utilization.gpu,
                'memory_utilization': utilization.memory
            })
        return gpu_usages
    
    def log_gpu_usage(self, epoch):
        gpu_usages = self.get_gpu_usage()
        usage_data = {
            'epoch': epoch,
            'gpu_utilization': gpu_usages[0]['gpu_utilization'],
            'memory_utilization': gpu_usages[0]['memory_utilization']
        }
        self.gpu_usage_data.append(usage_data)
        print(f"Epoch {epoch}: GPU {gpu_usages[0]['gpu']}: Memory Total: {gpu_usages[0]['memory_total']} | Memory Free: {gpu_usages[0]['memory_free']} | Memory Used: {gpu_usages[0]['memory_used']} | GPU Utilization: {gpu_usages[0]['gpu_utilization']}% | Memory Utilization: {gpu_usages[0]['memory_utilization']}%")
    
    def on_epoch_end(self, epoch, logs=None):
        self.log_gpu_usage(epoch+1)

    def on_train_end(self, logs=None):
        return self.gpu_usage_data

class ProfilerCallback(tf.keras.callbacks.Callback):
    def __init__(self, log_dir, log_frequency=3):
        super(ProfilerCallback, self).__init__()
        self.log_dir = log_dir
        self.log_frequency = log_frequency

    def on_epoch_begin(self, epoch, logs=None):
        if epoch % self.log_frequency == 0: 
            # Konfiguracja TensorFlow Profiler
            options = tf.profiler.experimental.ProfilerOptions(host_tracer_level=2) # Adjust CPU tracing level. Values are: 1 - critical info only, 2 - info, 3 - verbose. [default value is 2]
            profile_dir = os.path.join(self.log_dir, 'profiler')
            tf.profiler.experimental.start(profile_dir, options=options)
    
    def on_epoch_end(self, epoch, logs=None):
        if epoch % self.log_frequency == 0: 
            # Zatrzymanie TensorFlow Profiler
            tf.profiler.experimental.stop()

class MeasuringTime(tf.keras.callbacks.Callback):
    def __init__(self):
        super(MeasuringTime, self).__init__()
        self.start_time = 0
        self.total_time = 0

    def on_epoch_begin(self, epoch, logs=None):
        self.start_time = time.time()
    
    def on_epoch_end(self, epoch, logs=None):
        end_time = time.time()
        execution_time = end_time - self.start_time
        self.total_time += execution_time

    def on_train_end(self, logs=None):
        print("Training time: {}".format(self.total_time))
        return self.total_time