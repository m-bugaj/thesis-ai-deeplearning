# import os
# import tensorflow as tf
# from keras.models import Model
# from keras.layers import Flatten, Dense, Conv2D, MaxPooling2D, Dropout, Concatenate, Input, AveragePooling2D
# from keras.utils import to_categorical
# import cv2
# import numpy as np
# from PIL import Image

# class MnistClassifier:
#     def inception_module(self, x, filters):
#         # 1x1 conv
#         conv1 = Conv2D(filters[0], (1, 1), padding='same', activation='relu')(x)

#         # 1x1 conv -> 3x3 conv
#         conv3 = Conv2D(filters[1], (1, 1), padding='same', activation='relu')(x)
#         conv3 = Conv2D(filters[2], (3, 3), padding='same', activation='relu')(conv3)

#         # 1x1 conv -> 5x5 conv
#         conv5 = Conv2D(filters[3], (1, 1), padding='same', activation='relu')(x)
#         conv5 = Conv2D(filters[4], (5, 5), padding='same', activation='relu')(conv5)

#         # 3x3 max pooling -> 1x1 conv
#         pool = MaxPooling2D((3, 3), strides=(1, 1), padding='same')(x)
#         pool = Conv2D(filters[5], (1, 1), padding='same', activation='relu')(pool)

#         # concatenate filters
#         out = Concatenate()([conv1, conv3, conv5, pool])
#         return out

#     def load_png(self, data_path):
#         images = []
#         labels = []

#         for folder in os.listdir(data_path):
#             folder_path = os.path.join(data_path, folder)
#             for image_file in os.listdir(folder_path):
#                 img = Image.open(os.path.join(folder_path, image_file))
#                 img = img.convert('RGB')  # Ensure all images are RGB
#                 img = img.resize((224, 224))  # Resize image to 224x224 pixels
#                 img = np.array(img)  # Convert image to a numpy array
#                 img = img.astype('float16') / 255.0  # Normalize data to float values from 0 to 1
#                 images.append(img)

#                 label = folder
#                 labels.append(label)

#         return np.array(images), np.array(labels)

    
#     def train_model(self, model_name, compile_optimizer, compile_loss, fit_epochs, fit_batch_size):

#         input_layer = Input(shape=(224, 224, 3))

#         # Initial convolution and pooling layers
#         x = Conv2D(64, (7, 7), strides=(2, 2), padding='same', activation='relu')(input_layer)
#         x = MaxPooling2D((3, 3), strides=(2, 2), padding='same')(x)

#         # Inception modules
#         x = self.inception_module(x, [64, 96, 128, 16, 32, 32])
#         x = self.inception_module(x, [128, 128, 192, 32, 96, 64])
#         x = MaxPooling2D((3, 3), strides=(2, 2), padding='same')(x)

#         x = self.inception_module(x, [192, 96, 208, 16, 48, 64])
#         x = self.inception_module(x, [160, 112, 224, 24, 64, 64])
#         x = self.inception_module(x, [128, 128, 256, 24, 64, 64])
#         x = self.inception_module(x, [112, 144, 288, 32, 64, 64])
#         x = self.inception_module(x, [256, 160, 320, 32, 128, 128])
#         x = MaxPooling2D((3, 3), strides=(2, 2), padding='same')(x)

#         x = self.inception_module(x, [256, 160, 320, 32, 128, 128])
#         x = self.inception_module(x, [384, 192, 384, 48, 128, 128])

#         # Finishing layers
#         x = AveragePooling2D((7, 7), strides=(1, 1))(x)
#         x = Flatten()(x)
#         x = Dropout(0.4)(x)
#         x = Dense(10, activation='softmax')(x)  # 1000 classes in ImageNet

#         # Create model
#         model = Model(inputs=input_layer, outputs=x)
#         model.summary()

#         model.compile(optimizer=compile_optimizer, loss=compile_loss, metrics=['accuracy'])

#         x_train, y_train = self.load_png('C:\\Users\\BUGI\\Desktop\\Learn\\Magisterka\\DANE\\archive\\raw-img')
#         x_test, y_test = self.load_png('C:\\Users\\BUGI\\Desktop\\Learn\\Magisterka\\DANE\\archive\\test')

#         # Przekonwertowanie etykiety na kodowanie one-hot
#         y_train = to_categorical(y_train)
#         y_test = to_categorical(y_test)

#         model.fit(x_train, y_train, epochs=fit_epochs, batch_size=fit_batch_size)

#         # Generowanie ścieżki do pliku JSON
#         json_file_path = os.path.join('models', 'model' + model_name + '.json')
#         weights_file_path = os.path.join('models', 'model' + model_name + '.weights.h5')

#         # Zapisanie modelu do pliku JSON
#         json_string = model.to_json()
#         with open(json_file_path, 'w') as json_file:
#             json_file.write(json_string)

#         model.save_weights(weights_file_path)

#         score = model.evaluate(x_test, y_test, verbose=0)
#         print('Test loss:', score[0])
#         print('Test accuracy:', score[1])

import os
import tensorflow as tf
from keras.models import Model, load_model
from keras.layers import Flatten, Dense, Conv2D, MaxPooling2D, Dropout, Concatenate, Input, AveragePooling2D, GlobalAveragePooling2D
from keras.layers import BatchNormalization
from keras.layers.core import Activation
from keras.regularizers import l2
from keras.layers import concatenate
from keras.preprocessing.image import ImageDataGenerator
from keras.utils import to_categorical
from keras.optimizers import Adam
from keras import backend as K

import pandas as pd
import matplotlib.pyplot as plt
from pynvml import *
from keras.callbacks import TensorBoard, ModelCheckpoint
import datetime
import psutil
import time
from tensorflow import compat
from keras import backend as K
import shutil
import openpyxl

# class GoogLeNet:
#     @staticmethod
#     def conv_module(x, K, kX, kY, stride, chanDim,
#         padding = "same", reg = 0.0005, name = None):
#         # initialize the CONV, BN, and RELU layer names
#         (convName, bnName, actName) = (None, None, None)

#         # if a layer name was supplied, prepend it
#         if name is not None:
#             convName = name + "_conv"
#             bnName = name + "_bn"
#             actName = name + "_act"

#         # define a CONV => BN => RELU pattern
#         x = Conv2D(K, (kX, kY), strides = stride, padding = padding,
#             kernel_regularizer = l2(reg), name = convName)(x)
#         x = BatchNormalization(axis = chanDim, name = bnName)(x)
#         x = Activation("relu", name = actName)(x)

#         # return the block
#         return x
    
#     def inception_module(x, num1x1, num3x3Reduce, num3x3, num5x5Reduce,
#         num5x5, num1x1Proj, chanDim, stage, reg = 0.0005):
#         # define the first branch of the Inception module which
#         # consists of 1x1 convolutions
#         first = GoogLeNet.conv_module(x, num1x1, 1, 1, (1, 1),
#             chanDim, reg = reg, name = stage + "_first")

#         # define the second branch of the Inception module which
#         # consists of 1x1 and 3x3 convolutions
#         second = GoogLeNet.conv_module(x, num3x3Reduce, 1, 1, (1, 1),
#             chanDim, reg = reg, name = stage + "_second1")
#         second = GoogLeNet.conv_module(second, num3x3, 3, 3, (1, 1),
#             chanDim, reg = reg, name = stage + "_second2")

#         # define the third branch of the Inception module which
#         # are both 1x1 and 5x5 convolutions
#         third = GoogLeNet.conv_module(x, num5x5Reduce, 1, 1, (1, 1),
#             chanDim, reg = reg, name = stage + "_third1")
#         third = GoogLeNet.conv_module(third, num5x5, 5, 5, (1, 1),
#             chanDim, reg = reg, name = stage + "_third2")

#         # define the fourth branch of the Inception module which
#         # is the POOL projection
#         fourth = MaxPooling2D((3, 3), strides = (1, 1), padding = "same",
#             name = stage + "_pool")(x)
#         fourth = GoogLeNet.conv_module(fourth, num1x1Proj, 1, 1, (1, 1),
#             chanDim, reg = reg, name = stage + "_fourth")

#         # concatenate across the channel dimension
#         x = concatenate([first, second, third, fourth], axis = chanDim,
#             name = stage + "_mixed")

#         # return the block
#         return x
    
#     def build(width, height, depth, reg = 0.0005):
#         # initialize the input shape to be "channel last" and the
#         # channels dimension itself
#         inputShape = (height, width, depth)
#         chanDim = -1

#         # if we are using "channel first", update the input shape
#         # and channels dimension
#         if K.image_data_format() == "channels_first":
#             inputShape = (depth, height, width)
#             chanDim = 1
#         print(chanDim)
#         # define the model input, followed by a sequence of
#         # CONV => POOL => (CONV * 2) => POOL layers
#         inputs = Input(shape = inputShape)
#         x = GoogLeNet.conv_module(inputs, 64, 5, 5, (1, 1),
#             chanDim, reg = reg, name = "block1")
#         x = MaxPooling2D((3, 3), strides = (2, 2), padding = "same",
#             name = "pool1")(x)
#         x = GoogLeNet.conv_module(x, 64, 1, 1, (1, 1),
#             chanDim, reg = reg, name = "block2")
#         x = GoogLeNet.conv_module(x, 192, 3, 3, (1, 1),
#             chanDim, reg = reg, name = "block3")
#         x = MaxPooling2D((3, 3), strides = (2, 2), padding = "same",
#             name = "pool2")(x)

#         # apply two Inception module followed by a POOL
#         x = GoogLeNet.inception_module(x, 64, 96, 128, 16, 32, 32,
#             chanDim, "3a", reg = reg)
#         x = GoogLeNet.inception_module(x, 128, 128, 192, 32, 96, 64,
#             chanDim, "3b", reg = reg)
#         x = MaxPooling2D((3, 3), strides = (2, 2), padding = "same",
#             name = "pool3")(x)

#         # apply five Inception module followed by POOL
#         x = GoogLeNet.inception_module(x, 192, 96, 208, 16, 48, 64,
#             chanDim, "4a", reg = reg)
#         x = GoogLeNet.inception_module(x, 160, 112, 224, 24, 64, 64,
#             chanDim, "4b", reg = reg)
#         x = GoogLeNet.inception_module(x, 128, 128, 256, 24, 64, 64,
#             chanDim, "4c", reg = reg)
#         x = GoogLeNet.inception_module(x, 112, 144, 288, 32, 64, 64,
#             chanDim, "4d", reg = reg)
#         x = GoogLeNet.inception_module(x, 256, 160, 320, 32, 128, 128,
#             chanDim, "4e", reg = reg)
#         x = MaxPooling2D((3, 3), strides = (2, 2), padding = "same",
#             name = "pool4")(x)

#         # apply last two Inception module
#         x = GoogLeNet.inception_module(x, 256, 160, 320, 32, 128, 128,
#             chanDim, "5a", reg = reg)
#         x = GoogLeNet.inception_module(x, 384, 192, 384, 48, 128, 128,
#             chanDim, "5b", reg = reg)

#         # apply a POOL layer (average) followed by dropout
#         x = AveragePooling2D((4, 4), name = "pool5")(x)
#         x = Dropout(0.4, name = "do")(x)

#         # softmax classifier
#         x = Flatten(name = "flatten")(x)
#         x = Dense(10, kernel_regularizer = l2(reg), name = "labels")(x)
#         x = Activation("softmax", name = "softmax")(x)

#         # create the model
#         model = Model(inputs, x, name = "googlenet")

#         # return the constructed network architecture
#         return model

class GoogLeNet:
    # def __init__(self):
    #     # Inicjalizacja NVML
    #     nvmlInit()
    #     self.device_count = nvmlDeviceGetCount()
    #     self.gpu_usage_data = []
    
    # def __del__(self):
    #     # Zamykanie NVML przy zniszczeniu obiektu
    #     nvmlShutdown()

    def display_history(self, history, training_time, model_name, arch_name, accuracy, loss):
        # df = pd.DataFrame(history)
        # ax = df.plot(figsize=(8, 5))
        # plt.grid(True)
        # # plt.gca().set_ylim(0, 1)

        # # Dodanie tekstu pod wykresem
        # plt.subplots_adjust(bottom=0.2)  # Adjust bottom to make space for text
        # plt.text(0.5, -0.15, f'Czas treningu: {training_time:.2f} sekund', color='red', 
        #          ha='center', va='top', transform=ax.transAxes, fontsize=12)

        # plt.savefig('out/history.png')
        # plt.show()

        df = pd.DataFrame(history)
        
        fig, ax1 = plt.subplots(figsize=(8, 5))

        # Skala po lewej stronie dla accuracy
        ax1.plot(df.index, df['accuracy'], 'b-', label='Accuracy')
        ax1.plot(df.index, df['val_accuracy'], 'g-', label='Validation Accuracy')
        ax1.set_xlabel('Epoch')
        ax1.set_ylabel('Accuracy', color='b')
        ax1.tick_params(axis='y', labelcolor='b')
        ax1.legend(loc='upper left')

        parts = model_name.split('__')
        model_params = {}
        for part in parts:
            key_value = part.split('-', 1)  # Split only on the first dash
            key = key_value[0]
            value = key_value[1]
            if key in model_params:
                model_params[key] += '-' + value
            else:
                model_params[key] = value
        
        title_part = f"Batch size: {model_params['bs']}, Activation function: {model_params['af']}, Learning rate: {model_params['lr']}"

        ax1.set_title('Training History + GPU Usage:\n {} - {}'.format(arch_name, title_part))

        # Tworzenie drugiej osi y dla loss
        ax2 = ax1.twinx()
        ax2.plot(df.index, df['loss'], 'r-', label='Loss')
        ax2.plot(df.index, df['val_loss'], 'm-', label='Validation Loss')
        ax2.set_ylabel('Loss', color='r')
        ax2.tick_params(axis='y', labelcolor='r')
        ax2.legend(loc='upper right')

        fig.tight_layout()  # Dostosowanie layoutu, żeby elementy się nie nakładały
        ax1.grid(True)

        # Dodanie tekstu pod wykresem
        plt.subplots_adjust(bottom=0.25)  # Adjust bottom to make space for text
        plt.text(0.5, -0.30, f'Czas treningu: {training_time:.2f} sekund', color='red', 
                 ha='center', va='top', transform=ax1.transAxes, fontsize=12)
        
        # Wyświetlenie Accuracy i Loss
        plt.text(0.5, -0.16, f'Accuracy: {accuracy:.4f}', color='green', ha='center', va='top', transform=ax1.transAxes, fontsize=12)
        plt.text(0.5, -0.23, f'Loss: {loss:.4f}', color='blue', ha='center', va='top', transform=ax1.transAxes, fontsize=12)

        disp_dir = os.path.join("out", arch_name)
        if not os.path.exists(disp_dir):
            os.makedirs(disp_dir)
        plt.savefig(disp_dir + '/' + model_name + '__' + datetime.datetime.now().strftime("%Y%m%d-%H%M%S") + '.png')

        # plt.show()

    def display_combined_history(self, history, gpu_usage_data, training_time, model_name, arch_name, accuracy, loss):
        # Tworzenie wykresu łączonego dla historii trenowania i danych GPU
        history_df = pd.DataFrame(history)
        gpu_usage_df = pd.DataFrame(gpu_usage_data)
        
        fig, ax1 = plt.subplots(figsize=(8, 5))

        ax1.set_xlabel('Epoch')
        ax1.set_ylabel('Accuracy / Loss')
        ax1.plot(history_df.index, history_df['accuracy'], 'b-', label='Accuracy')
        ax1.plot(history_df.index, history_df['val_accuracy'], 'g-', label='Validation Accuracy')
        ax1.plot(history_df.index, history_df['loss'], 'r-', label='Loss')
        ax1.plot(history_df.index, history_df['val_loss'], 'm-', label='Validation Loss')
        ax1.legend(loc='upper left')

        parts = model_name.split('__')
        model_params = {}
        for part in parts:
            key_value = part.split('-', 1)  # Split only on the first dash
            key = key_value[0]
            value = key_value[1]
            if key in model_params:
                model_params[key] += '-' + value
            else:
                model_params[key] = value
        
        title_part = f"Batch size: {model_params['bs']}, Activation function: {model_params['af']}, Learning rate: {model_params['lr']}"

        ax1.set_title('Training History + GPU Usage:\n {} - {}'.format(arch_name, title_part))
        
        ax2 = ax1.twinx()
        ax2.set_ylabel('GPU Utilization (%) / Memory Utilization (%)')
        ax2.plot(gpu_usage_df.index, gpu_usage_df['gpu_utilization'], 'k-', label='GPU Utilization')
        ax2.plot(gpu_usage_df.index, gpu_usage_df['memory_utilization'], 'c-', label='Memory Utilization')
        ax2.legend(loc='upper right')

        fig.tight_layout()
        ax1.grid(True)
        plt.grid(True)
        # Dodanie tekstu pod wykresem
        plt.subplots_adjust(bottom=0.25)  # Adjust bottom to make space for text
        plt.text(0.5, -0.30, f'Czas treningu: {training_time:.2f} sekund', color='red', 
                 ha='center', va='top', transform=ax1.transAxes, fontsize=12)
        
        # Wyświetlenie Accuracy i Loss
        plt.text(0.5, -0.16, f'Accuracy: {accuracy:.4f}', color='green', ha='center', va='top', transform=ax1.transAxes, fontsize=12)
        plt.text(0.5, -0.23, f'Loss: {loss:.4f}', color='blue', ha='center', va='top', transform=ax1.transAxes, fontsize=12)

        disp_dir = os.path.join("out", arch_name)
        if not os.path.exists(disp_dir):
            os.makedirs(disp_dir)
        plt.savefig(disp_dir + '/' + 'combined__' + model_name + '__' + datetime.datetime.now().strftime("%Y%m%d-%H%M%S") + '.png')

        # plt.show()

    # Function to find the next available row in a given column
    def find_next_available_row(self, ws, column):
        row = 3
        while ws[f"{column}{row}"].value is not None:
            row += 1
        return row

    def inception_module(self, x, filters, activation_function):
        # 1x1 conv
        conv1 = Conv2D(filters[0], (1, 1), strides=(1, 1), padding='same', kernel_regularizer = l2(0.0002))(x)
        conv1 = BatchNormalization(axis = -1)(conv1)
        conv1 = Activation(activation_function)(conv1)
        # 1x1 conv -> 3x3 conv
        conv3 = Conv2D(filters[1], (1, 1), strides=(1, 1), padding='same', kernel_regularizer = l2(0.0002))(x)
        conv3 = BatchNormalization(axis = -1)(conv3)
        conv3 = Activation(activation_function)(conv3)
        conv3 = Conv2D(filters[2], (3, 3), strides=(1, 1), padding='same', kernel_regularizer = l2(0.0002))(conv3)
        conv3 = BatchNormalization(axis = -1)(conv3)
        conv3 = Activation(activation_function)(conv3)
        # 1x1 conv -> 5x5 conv
        conv5 = Conv2D(filters[3], (1, 1), strides=(1, 1), padding='same', kernel_regularizer = l2(0.0002))(x)
        conv5 = BatchNormalization(axis = -1)(conv5)
        conv5 = Activation(activation_function)(conv5)
        conv5 = Conv2D(filters[4], (5, 5), strides=(1, 1), padding='same', kernel_regularizer = l2(0.0002))(conv5)
        # 3x3 max pooling -> 1x1 conv
        pool = MaxPooling2D((3, 3), strides=(1, 1), padding='same')(x)
        pool = Conv2D(filters[5], (1, 1), strides=(1, 1), padding='same', kernel_regularizer = l2(0.0002))(pool)
        pool = BatchNormalization(axis = -1)(pool)
        pool = Activation(activation_function)(pool)
        # concatenate filters
        # out = Concatenate()([conv1, conv3, conv5, pool])
        out = concatenate([conv1, conv3, conv5, pool], axis=-1)
        return out
    
    def train_model(self, model_name, arch_name, compile_optimizer, compile_loss, fit_epochs, fit_batch_size, activation_function, log_custom_dir=''):
        input_layer = Input(shape=(224, 224, 3))
        # Initial convolution and pooling layers
        # x = Conv2D(64, (7, 7), strides=(2, 2), padding='same', activation='relu')(input_layer)
        # x = BatchNormalization(axis = -1)(x)
        # x = Activation("relu")(x)
        # x = MaxPooling2D((3, 3), strides=(2, 2), padding='same')(x)

        x = Conv2D(64, (5, 5), strides=(1, 1), padding='same', kernel_regularizer = l2(0.0002))(input_layer)
        x = BatchNormalization(axis = -1)(x)
        x = Activation(activation_function)(x)
        x = MaxPooling2D((3, 3), strides=(2, 2), padding='same')(x)

        x = Conv2D(64, (1, 1), strides=(1, 1), padding='same', kernel_regularizer = l2(0.0002))(x)
        x = BatchNormalization(axis = -1)(x)
        x = Activation(activation_function)(x)
        x = Conv2D(192, (3, 3), strides=(1, 1), padding='same', kernel_regularizer = l2(0.0002))(x)
        x = BatchNormalization(axis = -1)(x)
        x = Activation(activation_function)(x)
        x = MaxPooling2D((3, 3), strides=(2, 2), padding='same')(x)

        # Inception modules
        x = self.inception_module(x, [64, 96, 128, 16, 32, 32], activation_function)
        x = self.inception_module(x, [128, 128, 192, 32, 96, 64], activation_function)
        x = MaxPooling2D((3, 3), strides=(2, 2), padding='same')(x)
        x = self.inception_module(x, [192, 96, 208, 16, 48, 64], activation_function)
        x = self.inception_module(x, [160, 112, 224, 24, 64, 64], activation_function)
        x = self.inception_module(x, [128, 128, 256, 24, 64, 64], activation_function)
        x = self.inception_module(x, [112, 144, 288, 32, 64, 64], activation_function)
        x = self.inception_module(x, [256, 160, 320, 32, 128, 128], activation_function)
        x = MaxPooling2D((3, 3), strides=(2, 2), padding='same')(x)
        x = self.inception_module(x, [256, 160, 320, 32, 128, 128], activation_function)
        x = self.inception_module(x, [384, 192, 384, 48, 128, 128], activation_function)
        # Finishing layers
        # x = AveragePooling2D((7, 7), strides=(1, 1))(x)
        x = AveragePooling2D((4, 4))(x)
        # x = GlobalAveragePooling2D()(x)
        # x = AveragePooling2D((7, 7))(x)
        x = Dropout(0.4)(x)
        x = Flatten()(x)
        # x = Dense(10, activation='softmax')(x)  # Assuming 10 classes for MNIST-like example
        x = Dense(10, kernel_regularizer = l2(0.0002))(x)
        x = Activation("softmax")(x)

        # Create model
        model = Model(inputs=input_layer, outputs=x)

        
        model.summary()

        print("Num GPUs Available: ", len(tf.config.experimental.list_physical_devices('GPU')))

        # reduce_lr = tf.keras.callbacks.ReduceLROnPlateau(monitor='val_loss', factor=0.2, patience=5, min_lr=0.0001)
        # callbacks = [reduce_lr]


        # model = GoogLeNet.build(width = 224, height = 224, depth = 3, reg = 0.0002)

        model.compile(optimizer=compile_optimizer, loss=compile_loss, metrics=['accuracy'])

        train_datagen = ImageDataGenerator(rescale=1./255)
        test_datagen = ImageDataGenerator(rescale=1./255)

        train_generator = train_datagen.flow_from_directory(
            'DANE/etap1/raw-img',
            target_size=(224, 224),
            batch_size=fit_batch_size,
            class_mode='categorical')

        test_generator = test_datagen.flow_from_directory(
            'DANE/etap1/test',
            target_size=(224, 224),
            batch_size=fit_batch_size,
            class_mode='categorical')

        # Konfiguracja TensorBoard
        log_dir = os.path.join(log_custom_dir, "logs", "fit", arch_name, model_name + '__' + datetime.datetime.now().strftime("%Y%m%d-%H%M%S"))
        tensorboard_callback = TensorBoard(log_dir=log_dir, histogram_freq=4)

        # Callbacks definitions
        system_usage_logger = SystemUsageLogger(log_dir=log_dir, log_frequency=4)
        image_callback = TensorBoardImageCallback(log_dir, train_generator, test_generator, log_frequency=4)
        log_gpu_usage_callback = GPUUsageLogger()
        profiler_callback = ProfilerCallback(log_dir=log_dir, log_frequency=8)
        measuring_time = MeasuringTime()

        model_file_path = os.path.join('models', arch_name)
        if not os.path.exists(model_file_path):
            os.makedirs(model_file_path)
        ckpt_file_path_with_name = os.path.join(model_file_path, 'model__' + model_name + '__' + datetime.datetime.now().strftime("%Y%m%d-%H%M%S") + '.ckpt')
        model_checkpoint_callback = ModelCheckpoint(
            filepath=ckpt_file_path_with_name,  # Ścieżka do pliku zapisu modelu
            save_best_only=True,       # Zapisuj tylko najlepszy model
            monitor='val_accuracy',    # Monitorowana wartość (dokładność walidacyjna)
            mode='max',                # Tryb maksymalizacji wartości (dla 'accuracy')
            verbose=1                  # Wyświetl komunikaty o zapisie modelu
        )


        start_time = time.time()

        history = model.fit(train_generator, 
                                steps_per_epoch=train_generator.samples // fit_batch_size, 
                                epochs=fit_epochs, 
                                validation_data=test_generator, 
                                callbacks = [measuring_time, 
                                             log_gpu_usage_callback, 
                                             system_usage_logger, 
                                             tensorboard_callback, 
                                             image_callback, 
                                             profiler_callback,
                                             model_checkpoint_callback])
        
        stop_time = time.time()

        # Pobieranie danych z `gpu_logger` po zakończeniu trenowania
        gpu_usage_data = log_gpu_usage_callback.on_train_end()

        training_time = measuring_time.on_train_end()

        model = load_model(ckpt_file_path_with_name)
        model_file_path_with_name = os.path.join(model_file_path, 'model__' + model_name + '__' + datetime.datetime.now().strftime("%Y%m%d-%H%M%S") + '.keras')
        model.save(model_file_path_with_name)
        model = load_model(model_file_path_with_name)

        score = model.evaluate(test_generator, verbose=0)
        print('Test loss:', score[0])
        print('Test accuracy:', score[1])
        print("Total time: {}".format(stop_time-start_time))

        # Obliczanie średnich wartości zużycia GPU oraz pamięci
        total_gpu_utilization = sum(data['gpu_utilization'] for data in gpu_usage_data)
        total_memory_utilization = sum(data['memory_utilization'] for data in gpu_usage_data)
        data_len = len(gpu_usage_data)

        average_gpu_utilization = total_gpu_utilization / data_len if data_len > 0 else 0
        average_memory_utilization = total_memory_utilization / data_len if data_len > 0 else 0

        print(f"Average GPU Utilization: {average_gpu_utilization}%")
        print(f"Average Memory Utilization: {average_memory_utilization}%")

        # Obliczanie średnich wartości zużycia RAM oraz CPU
        used_ram_data = system_usage_logger.used_ram_data
        cpu_usage_data = system_usage_logger.cpu_usage_data

        average_used_ram = sum(used_ram_data) / len(used_ram_data) if used_ram_data else 0
        average_cpu_usage = sum(cpu_usage_data) / len(cpu_usage_data) if cpu_usage_data else 0

        print(f"Average Used RAM: {average_used_ram}%")
        print(f"Average CPU Usage: {average_cpu_usage}%")

        # Path to the Excel file
        excel_file_path = os.path.join("out", arch_name, 'Excels')
        if not os.path.exists(excel_file_path):
            os.makedirs(excel_file_path)

        excel_file_path_with_name = os.path.join(excel_file_path, 'excel_name' + '.xlsx')

        # Load or create a new Excel file
        if os.path.exists(excel_file_path_with_name):
            wb = openpyxl.load_workbook(excel_file_path_with_name)
        else:
            wb = openpyxl.Workbook()

        # Select the active sheet
        ws = wb.active

        # Find the next available rows for each column
        row_accuracy = self.find_next_available_row(ws, 'B')
        row_loss = self.find_next_available_row(ws, 'C')
        row_time = self.find_next_available_row(ws, 'D')
        row_gpu_util = self.find_next_available_row(ws, 'E')
        row_mem_util = self.find_next_available_row(ws, 'F')
        row_ram_usage = self.find_next_available_row(ws, 'G')
        row_cpu_usage = self.find_next_available_row(ws, 'H')

        # Ensure rows align (take the maximum row number to avoid overwriting)
        next_row = max(row_accuracy, row_loss, row_time, row_gpu_util, row_mem_util, row_ram_usage, row_cpu_usage)

        # Write the data to the next available row
        ws[f"B{next_row}"] = score[1]
        ws[f"C{next_row}"] = score[0]
        ws[f"D{next_row}"] = training_time
        ws[f"E{next_row}"] = average_gpu_utilization
        ws[f"F{next_row}"] = average_memory_utilization
        ws[f"G{next_row}"] = average_used_ram
        ws[f"H{next_row}"] = average_cpu_usage

        # ws[f"C{next_row}"] = f"{score[0]:.4f}".replace(',', '.')
        # ws[f"D{next_row}"] = f"{training_time:.2f}".replace(',', '.')
        # ws[f"E{next_row}"] = f"{average_gpu_utilization:.2f}".replace(',', '.')
        # ws[f"F{next_row}"] = f"{average_memory_utilization:.2f}".replace(',', '.')
        # ws[f"G{next_row}"] = f"{average_used_ram:.2f}".replace(',', '.')
        # ws[f"H{next_row}"] = f"{average_cpu_usage:.2f}".replace(',', '.')

        # Add the formula in column I starting from row 3 down to the current row
        for row in range(3, next_row + 1):
            formula = f"=ABS(- (3 * (1 - B{row})) - (2 * C{row}) - (2 * (D{row} / MAX(D$3:D$1048576))) - (1 * (E{row} / 100)) - (1 * (F{row} / 100)) - (1 * (G{row} / 100)) - (1 * (H{row} / 100)))"
            ws[f"I{row}"].value = formula
        # Save the workbook
        wb.save(excel_file_path_with_name)

        # Wyświetlenie historii trenowania oraz danych dotyczących użycia GPU
        self.display_history(history.history, training_time, model_name, arch_name, score[1], score[0])
        self.display_combined_history(history.history, gpu_usage_data, training_time, model_name, arch_name, score[1], score[0])

        # Czyszczenie sesji Keras, aby zwolnić pamięć
        K.clear_session()
        compat.v1.reset_default_graph()

        shutil.rmtree(ckpt_file_path_with_name)

        del model, score, history, train_datagen, test_datagen, train_generator, test_generator, model_checkpoint_callback, compile_optimizer



class TensorBoardImageCallback(tf.keras.callbacks.Callback):
    def __init__(self, log_dir, train_data, test_data, log_frequency=1):
        super().__init__()
        self.log_dir = log_dir
        self.train_data = train_data
        self.test_data = test_data
        self.writer = tf.summary.create_file_writer(os.path.join(log_dir, 'images'))
        self.log_frequency = log_frequency

    def on_epoch_end(self, epoch, logs=None):
        if epoch % self.log_frequency == 0:
            # Wybierz kilka przykładów z danych treningowych i walidacyjnych
            train_images, _ = next(self.train_data)
            test_images, _ = next(self.test_data)
            
            # Przygotuj obrazy do zapisu
            with self.writer.as_default():
                tf.summary.image('Training Images', train_images, step=epoch)
                tf.summary.image('Test Images', test_images, step=epoch)

class SystemUsageLogger(tf.keras.callbacks.Callback):
    def __init__(self, log_dir, log_frequency=3):
        super(SystemUsageLogger, self).__init__()
        self.log_dir = log_dir
        self.file_writer = tf.summary.create_file_writer(log_dir)
        self.log_frequency = log_frequency
        self.used_ram_data = []
        self.cpu_usage_data = []
    
    def on_epoch_end(self, epoch, logs=None):
        if epoch % self.log_frequency == 0:    
            # Log system usage
            memory_info = psutil.virtual_memory()
            used_ram = memory_info.used / (1024 ** 3)  # Convert to GB
            used_ram_percent = memory_info.percent
            available_ram = memory_info.available / (1024 ** 3)  # Convert to GB
            cpu_usage = psutil.cpu_percent()

            self.used_ram_data.append(used_ram_percent)
            self.cpu_usage_data.append(cpu_usage)

            with self.file_writer.as_default():
                tf.summary.scalar('used_ram', used_ram, step=epoch)
                tf.summary.scalar('available_ram', available_ram, step=epoch)
                tf.summary.scalar('cpu_usage', cpu_usage, step=epoch)

class GPUUsageLogger(tf.keras.callbacks.Callback):
    def __init__(self):
        super(GPUUsageLogger, self).__init__()
        nvmlInit()
        self.device_count = nvmlDeviceGetCount()
        self.gpu_usage_data = []

    def __del__(self):
        # Zamykanie NVML przy zniszczeniu obiektu
        nvmlShutdown()

    def get_gpu_usage(self):
        gpu_usages = []
        for i in range(self.device_count):
            handle = nvmlDeviceGetHandleByIndex(i)
            info = nvmlDeviceGetMemoryInfo(handle)
            utilization = nvmlDeviceGetUtilizationRates(handle)
            gpu_usages.append({
                'gpu': i,
                'memory_total': info.total,
                'memory_free': info.free,
                'memory_used': info.used,
                'gpu_utilization': utilization.gpu,
                'memory_utilization': utilization.memory
            })
        return gpu_usages
    
    def log_gpu_usage(self, epoch):
        gpu_usages = self.get_gpu_usage()
        usage_data = {
            'epoch': epoch,
            'gpu_utilization': gpu_usages[0]['gpu_utilization'],
            'memory_utilization': gpu_usages[0]['memory_utilization']
        }
        self.gpu_usage_data.append(usage_data)
        print(f"Epoch {epoch}: GPU {gpu_usages[0]['gpu']}: Memory Total: {gpu_usages[0]['memory_total']} | Memory Free: {gpu_usages[0]['memory_free']} | Memory Used: {gpu_usages[0]['memory_used']} | GPU Utilization: {gpu_usages[0]['gpu_utilization']}% | Memory Utilization: {gpu_usages[0]['memory_utilization']}%")
    
    def on_epoch_end(self, epoch, logs=None):
        self.log_gpu_usage(epoch+1)

    def on_train_end(self, logs=None):
        return self.gpu_usage_data

class ProfilerCallback(tf.keras.callbacks.Callback):
    def __init__(self, log_dir, log_frequency=3):
        super(ProfilerCallback, self).__init__()
        self.log_dir = log_dir
        self.log_frequency = log_frequency

    def on_epoch_begin(self, epoch, logs=None):
        if epoch % self.log_frequency == 0: 
            # Konfiguracja TensorFlow Profiler
            options = tf.profiler.experimental.ProfilerOptions(host_tracer_level=2) # Adjust CPU tracing level. Values are: 1 - critical info only, 2 - info, 3 - verbose. [default value is 2]
            profile_dir = os.path.join(self.log_dir, 'profiler')
            tf.profiler.experimental.start(profile_dir, options=options)
    
    def on_epoch_end(self, epoch, logs=None):
        if epoch % self.log_frequency == 0: 
            # Zatrzymanie TensorFlow Profiler
            tf.profiler.experimental.stop()

class MeasuringTime(tf.keras.callbacks.Callback):
    def __init__(self):
        super(MeasuringTime, self).__init__()
        self.start_time = 0
        self.total_time = 0

    def on_epoch_begin(self, epoch, logs=None):
        self.start_time = time.time()
    
    def on_epoch_end(self, epoch, logs=None):
        end_time = time.time()
        execution_time = end_time - self.start_time
        self.total_time += execution_time

    def on_train_end(self, logs=None):
        print("Training time: {}".format(self.total_time))
        return self.total_time